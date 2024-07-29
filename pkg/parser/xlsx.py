"""
This file contains functions to handle Excel file operations such as reading
and covertiing it into a csv files for further processing. These functions can be used to read
Excel files, convert them to csv files, and extract data for analysis.
"""

import openpyxl
import os
import csv
from config.config_loader import Config
from pkg.db.db import FileManager
from pkg.db.progressEnum import ProgressEnum
from pkg.db.csv_type import CSVType
from pkg.hashing.hash import make_hash
AppConfig = Config.load_config()


def read_file(path, file_name):
    """
    read_file
    This function will read from the given file path and file name.
    Args:
    path: The path to the file.
    file_name: The name of the file.
    """
    final_path = f"{AppConfig["app"]["downloadDirectory"]}/{path}/{file_name}"
    try:
        wb = openpyxl.load_workbook(final_path)
        sheet = wb.active
        for row in sheet.iter_rows():
            for cell in row:
                print(cell.value)
    except Exception as e:
        print(f"Error occurred: {e}")
        raise
    return True

def convert_to_csv(path, file_name):
    """
    convert_to_csv
    This function will convert the given Excel file to a CSV file.
    Args:
    path: The path to the file.
    file_name: The name of the file.
    """
    final_path = f"{AppConfig["app"]["downloadDirectory"]}/{path}/{file_name}"
    print(f"Converting {final_path} to CSV...")
    os.makedirs(final_path.split('.')[0], exist_ok=True)
    try:
        wb = openpyxl.load_workbook(final_path,data_only=True)
        fm = FileManager()
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # Construct a unique CSV file name for each sheet
            csv_file_path = f"{final_path.split('.')[0]}/{sheet_name}.csv"
            with open(csv_file_path, "w", encoding='utf-8',newline='') as csv_file:
                writer = csv.writer(csv_file)
                for row in sheet.iter_rows(values_only=True):  # Use values_only=True to get calculated values
                    writer.writerow([str(cell) if cell is not None else '' for cell in row])
                # print(f"CSV file {csv_file.read()} created successfully")
            with open(csv_file_path, "rb") as csv_file:
                current_hash = make_hash(csv_file.read())
            
            print(
                (path+"/"+file_name).split('.')[0],
                csv_file_path.split('/')[-1],
                current_hash,
                ProgressEnum.READYTOPARSE,
                "csv"," has been created successfully."
            )
            existing_file = fm.read_file_record_by_path_and_filename(
                (path+"/"+file_name).split('.')[0], csv_file_path.split('/')[-1]
            )
            if existing_file is not None:
                fm.update_file_record(
                    existing_file.id,
                    sha256=current_hash,
                    progress=ProgressEnum.READYTOPARSE,
                    filetype="csv",
                )
            else:
                fm.create_file_record(
                    filename=csv_file_path.split('/')[-1],
                    path=(path+"/"+file_name).split('.')[0],
                    sha256=current_hash,
                    progress=ProgressEnum.READYTOPARSE,
                    filetype="csv",
                )
    except Exception as e:
        print(f"Error occurred: {e}")
        raise
    return True

def find_csv_type(path, file_name):
    """
    find_csv_type
    This function will find the type of the CSV file.
    Args:
    path: The path to the file.
    file_name: The name of the file.
    """
    fm=FileManager()
    try:
        # print(f"Finding CSV type for {path.split('/')[-1]}{file_name}...")
        primary_filename = path.split('/')[-1]
        if primary_filename == "Payroll Validation":
            existing_file = fm.read_file_record_by_path_and_filename(path, file_name)
            print(existing_file)
            if existing_file.csv_type== CSVType.OUTPUT: return True
            fm.update_file_record_from_path_name(
                path=path,
                filename=file_name,
                csv_type=CSVType.PAYROLL
            )
        if primary_filename == "Vendor Document Checklist":
            existing_file = fm.read_file_record_by_path_and_filename(path, file_name)
            print(existing_file)
            if existing_file.csv_type== CSVType.OUTPUT: return True            
            fm.update_file_record_from_path_name(
                path=path,
                filename=file_name,
                csv_type=CSVType.VENDOR
            )
        if primary_filename == "PO and Invoice Workflow data":
            existing_file = fm.read_file_record_by_path_and_filename(path, file_name)
            print(existing_file)
            if existing_file.csv_type== CSVType.OUTPUT: return True
            if file_name == "PO.csv":
                fm.update_file_record_from_path_name(
                    path=path,
                    filename=file_name,
                    csv_type=CSVType.PO
                )
            if file_name == "Invoice.csv":
                fm.update_file_record_from_path_name(
                    path=path,
                    filename=file_name,
                    csv_type=CSVType.INVOICE
                )
    except Exception as e:
        print(f"Error occurred: {e}")
        raise
    return True

def test():
    fm=FileManager()
    try:

        all_xlsx_files = fm.read_all_files_by_filetype("xlsx")
        print("All xlsx files:", all_xlsx_files)
        for i in all_xlsx_files:
            convert_to_csv(i.path, i.filename)
            print(f"{i.path}\t{i.filename}\t{i.csv_type}")
        all_csv_files = fm.read_all_files_by_filetype("csv")
        print(all_csv_files)
        for i in all_csv_files:
            find_csv_type(i.path, i.filename)
            print(f"{i.path}\t{i.filename}\t{i.csv_type}")
    except:
        print("Error occurred")
        raise
if __name__ == "__main__":
    read_file("/sites/NAF/Shared Documents/","Book.xlsx")
